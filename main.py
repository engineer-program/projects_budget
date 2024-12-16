from datetime import date
import os
from openpyxl import load_workbook, Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, Alignment, PatternFill
from tkinter import *
from tkinter import messagebox, ttk
from src.budget_projects_creater.employees import *
from src.budget_projects_creater.working_hours import *


TYPES_FINANCES = ["З/П", 
                  "Премия"]

BASIC_FONT = Font(bold=True, 
                 name='Arial', 
                 size=10)
CENTER_ALIGNMENT_WITH_WRAP = Alignment(horizontal='center', 
                                       vertical='center', 
                                       wrap_text=True)
CENTER_ALIGNMENT_WITHOUT_WRAP = Alignment(horizontal='center', 
                                          vertical='center', 
                                          wrap_text=False)


def full_sheet_title(sheet_name: str, year: float = date.today().year) -> str:
    return f'{sheet_name} за {str(year)} год'


def num_col_to_char_excel(col: float) -> str:
    integer_part = col // 27
    fract_part = col % 27
    res = ''
    if integer_part > 0:
        res += chr(64+integer_part)
        res += chr(64+fract_part+1)
    else:
        res += chr(64+fract_part)

    return res


def row_col_to_cell_excel(row: int,
                          col: int) -> str:
    return(num_col_to_char_excel(col)+str(row))


def parse_column_names_projects_to_list(filepath: str) -> list:
    if not os.path.exists(filepath):
        FileExistsError(f'Файл отсутствует!')

    wb = load_workbook(filepath)
    sheet = wb.worksheets[0]

    first_col = 1
    while True:
        if 'Номер' in sheet.cell(4, first_col).value:
            break
        first_col += 1

    column_names_projects = []
    for col in range(first_col, first_col+4):
        column_names_projects.append(sheet.cell(4, col).value)

    wb.close()
    return column_names_projects
    

def parse_info_employees_projects_to_lists(filepath: str, 
                                           year: float = date.today().year) -> list:
    if not os.path.exists(filepath):
        FileExistsError(f'Файл отсутствует!')

    wb = load_workbook(filepath)
    sheet = wb.worksheets[0]

    col_num_project = 1
    while True:
        if 'Номер' in sheet.cell(4, col_num_project).value:
            break
        col_num_project += 1
    
    col_position = 3
    col_date = 9
    col_hours = 10
    employees = {}
    projects = {}
    row = 5
    while True:
        user = sheet.cell(row, 1).value
        if user is None:
            break
        patronymic = sheet.cell(row, 2).value
        employee = Employee(*user.split(), patronymic)

        if employee.get_full_name() in employees.keys():
            employee = employees[employee.get_full_name()]

        date = datetime.strptime(sheet.cell(row, col_date).value, '%Y-%m-%d')
        if year == str(date.year):
            employee.set_position(sheet.cell(row, col_position).value)
            num_proj = str(sheet.cell(row, col_num_project).value).strip()
            hours = sheet.cell(row, col_hours).value
            employee.add_work_hours(date, num_proj, hours)
            employees[employee.get_full_name()] = employee

            if num_proj not in projects:
                projects[num_proj] = [sheet.cell(row, col_num_project+i).value for i in range(1, 4)]
                if str(num_proj).lower() == 'проработка':
                    projects[num_proj][0] = 'Проработка'
        row += 1

    wb.close()
    return list(employees.values()), dict(sorted(projects.items()))


def create_empty_budget_projects_sheet(filename: str, 
                                 project_names_list: list, 
                                 column_names: list, 
                                 year: float = date.today().year, 
                                 sheet_name: str = "Бюджеты проектов"):
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()

    sheet_title = full_sheet_title(sheet_name, year)
    if sheet_title not in wb.sheetnames:
        sheet = wb.create_sheet(sheet_title)
        sheet = wb[sheet_title]

        column_names.append(f'Бюджет на 01.01.{year} г.')
        for month in MONTHS:
            column_names.append(f'{month} - Премия')

        sheet.freeze_panes = 'D2'
        columns_width = [20, 40, 30, 30]
        for col in range(len(column_names)):
            sheet.cell(1, col+1).value = column_names[col]
            if col < len(columns_width):
                sheet.column_dimensions[num_col_to_char_excel(col+1)].width = columns_width[col]
                sheet.column_dimensions[num_col_to_char_excel(col+1)].font = BASIC_FONT
            else:
                sheet.column_dimensions[num_col_to_char_excel(col+1)].width = len(column_names[col]) + 1
                sheet.column_dimensions[num_col_to_char_excel(col+1)].alignment = CENTER_ALIGNMENT_WITHOUT_WRAP
        sheet.row_dimensions[1].font = BASIC_FONT

        row = 2
        for project in project_names_list:
            for i in range(len(project)):
                sheet.cell(row, i+1).value = project[i]
                sheet.cell(row, i+1).alignment = CENTER_ALIGNMENT_WITH_WRAP
            row += 1
    else:
        messagebox.showerror("Ошибка", f'Лист "{sheet_title}" уже существует в файле "{filename}"')
        raise ValueError('Лист уже существует в файле!')
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    wb.save(filename)
    wb.close()
    

def create_empty_finance_employees_sheet(filename: str, 
                                   employee_position_list: list, 
                                   year: float = date.today().year, 
                                   sheet_name: str = "Финансы сотрудников"):
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()

    sheet_title = full_sheet_title(sheet_name, year)
    if sheet_title not in wb.sheetnames:
        sheet = wb.create_sheet(sheet_title)
        sheet = wb[sheet_title]
        sheet.freeze_panes = 'C2'

        sheet.cell(1, 1).value = "ФИО сотрудника"
        sheet.cell(1, 1).alignment = CENTER_ALIGNMENT_WITHOUT_WRAP
        sheet.cell(1, 2).value = "Должность"
        sheet.cell(1, 2).alignment = CENTER_ALIGNMENT_WITHOUT_WRAP
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 60

        col = 3
        for month in MONTHS:
            for income in TYPES_FINANCES:
                sheet.cell(1, col).value = f'{month} - {income}'
                sheet.column_dimensions[num_col_to_char_excel(col)].width = len(sheet.cell(1, col).value) + 1
                sheet.column_dimensions[num_col_to_char_excel(col)].alignment = CENTER_ALIGNMENT_WITHOUT_WRAP
                col += 1

        sheet.row_dimensions[1].font = BASIC_FONT
        sheet.column_dimensions['A'].font = BASIC_FONT
        sheet.column_dimensions['A'].alignment = CENTER_ALIGNMENT_WITH_WRAP
        sheet.column_dimensions['B'].font = BASIC_FONT

        row = 2
        for employee, position in employee_position_list:
            sheet.cell(row, 1).value = employee
            sheet.cell(row, 2).value = position
            sheet.cell(row, 2).alignment = CENTER_ALIGNMENT_WITH_WRAP
            row += 1
    else:
        messagebox.showerror("Ошибка", f'Лист "{sheet_title}" уже существует в файле "{filename}"')
        raise ValueError('Лист уже существует в файле!')
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    wb.save(filename)
    wb.close()


def add_new_employees(filename: str, 
                      employee_position_list: list, 
                      year: float = date.today().year, 
                      sheet_name: str = "Финансы сотрудников"):
    
    sheet_title = full_sheet_title(sheet_name, year)
    if not os.path.exists(filename) or sheet_title not in load_workbook(filename).sheetnames:
        create_empty_finance_employees_sheet(filename, employee_position_list, year, sheet_name)
    else:
        wb = load_workbook(filename)
        sheet = wb[sheet_title]
        sheet.freeze_panes = 'C2'
        row = 2
        for employee, position in employee_position_list:
            employee_lower = employee.strip().lower()
            while True:
                employee_excel = sheet.cell(row, 1).value
                if employee_excel is not None:
                    employee_excel = str(employee_excel).strip().lower()
                    if employee_lower != employee_excel:
                        if employee_lower == sorted([employee_lower, employee_excel])[0]:
                            sheet.insert_rows(row)
                            sheet.cell(row, 1).value = employee
                            sheet.cell(row, 2).value = position
                            sheet.cell(row, 2).alignment = CENTER_ALIGNMENT_WITH_WRAP
                            for col in range(1, 25):
                                sheet.cell(row, 2+col).value = None
                            row += 1
                            break
                        else:
                            row += 1
                    else:
                        row += 1
                        break
                else:
                    sheet.cell(row, 1).value = employee
                    sheet.cell(row, 2).value = position
                    sheet.cell(row, 2).alignment = CENTER_ALIGNMENT_WITH_WRAP
                    for col in range(1, 25):
                        sheet.cell(row, 2+col).value = None
                    row += 1
                    break
        
        wb.save(filename)
        wb.close()


def add_new_projects(filename: str, 
                     project_names_list: list, 
                     column_names: list, 
                     year: float = date.today().year, 
                     sheet_name: str = "Бюджеты проектов"):
    
    sheet_title = full_sheet_title(sheet_name, year)
    if not os.path.exists(filename) or sheet_title not in load_workbook(filename).sheetnames:
        create_empty_budget_projects_sheet(filename, project_names_list, column_names, year, sheet_name)
    else:
        wb = load_workbook(filename)
        sheet = wb[sheet_title]
        sheet.freeze_panes = 'D2'
        row = 2
        for num_project, *names_project in project_names_list:
            num_project_lower = num_project.lower()
            while True:
                num_project_excel = sheet.cell(row, 1).value
                if num_project_excel is not None:
                    num_project_excel = str(num_project_excel).strip().lower()
                    if num_project_lower != num_project_excel:
                        if num_project_lower == sorted([num_project_lower, num_project_excel])[0]:
                            sheet.insert_rows(row)
                            sheet.cell(row, 1).value = num_project
                            sheet.cell(row, 1).alignment = CENTER_ALIGNMENT_WITH_WRAP
                            for i in range(len(names_project)):
                                sheet.cell(row, 2+i).value = names_project[i]
                                sheet.cell(row, 2+i).alignment = CENTER_ALIGNMENT_WITH_WRAP
                            for col in range(1, 20):
                                sheet.cell(row, 2+len(names_project)+col).value = None
                            row += 1
                            break
                        else:
                            row += 1
                    else:
                        row += 1
                        break
                else:
                    sheet.cell(row, 1).value = num_project
                    sheet.cell(row, 1).alignment = CENTER_ALIGNMENT_WITH_WRAP
                    for i in range(len(names_project)):
                        sheet.cell(row, 2+i).value = names_project[i]
                        sheet.cell(row, 2+i).alignment = CENTER_ALIGNMENT_WITH_WRAP
                    for col in range(1, 20):
                        sheet.cell(row, 2+len(names_project)+col).value = None
                    row += 1
                    break
        
        wb.save(filename)
        wb.close()


def parse_completed_finance_employees(filename: str, 
                                      year: float = date.today().year, 
                                      sheet_name: str = "Финансы сотрудников") -> dict:
    
    sheet_title = full_sheet_title(sheet_name, year)
    if not os.path.exists(filename) or sheet_title not in load_workbook(filename).sheetnames:
        messagebox.showerror('Ошибка', 'Неверное имя файла или отсутствует необходимый лист в файле!')
        raise NameError('Неверное имя файла или отсутствует необходимый лист в файле!')

    wb = load_workbook(filename)
    sheet = wb[sheet_title]
    finance_employees = {}

    row = 2
    while True:
        employee = sheet.cell(row, 1).value
        if employee is None:
            break
        month_finance = {}

        col = 3
        while True:
            type_income = sheet.cell(1, col).value
            if type_income is None:
                break
            month, type_finance = [x.strip() for x in type_income.split('-')]
            month_finance[type_finance] = month_finance.get(type_finance, {})
            month_finance[type_finance][month] = sheet.cell(row, col).value
            col += 1
        finance_employees[employee.strip()] = month_finance
        row += 1

    wb.close()
    
    return finance_employees


def parse_completed_finance_projects(filename: str, 
                                     year: float = date.today().year, 
                                     sheet_name: str = "Бюджеты проектов") -> dict:
    
    sheet_title = full_sheet_title(sheet_name, year)
    if not os.path.exists(filename) or sheet_title not in load_workbook(filename).sheetnames:
        messagebox.showerror("Ошибка", 'Неверное имя файла или отсутствует необходимый лист в файле!')
        return None

    wb = load_workbook(filename)
    sheet = wb[sheet_title]
    finance_projects = {}

    row = 2
    while True:
        project = sheet.cell(row, 1).value
        if project is None:
            break
        incomes_project = {}

        col = 5
        while True:
            income = sheet.cell(1, col).value
            if income is None:
                break
            incomes_project[income.strip()] = sheet.cell(row, col).value
            col += 1
        finance_projects[str(project).strip()] = incomes_project
        row += 1

    wb.close()

    return finance_projects


def parse_all_projects(filename: str, 
                        year: float = date.today().year, 
                        sheet_name: str = "Бюджеты проектов") -> dict:
    sheet_title = full_sheet_title(sheet_name, year)
    if not os.path.exists(filename) or sheet_title not in load_workbook(filename).sheetnames:
        messagebox.showerror("Ошибка", 'Неверное имя файла или отсутствует необходимый лист в файле!')
        return None

    wb = load_workbook(filename)
    sheet = wb[sheet_title]
    all_projects_dict = {}
    row = 2
    while True:
        project = sheet.cell(row, 1).value
        if project is None:
            break
        all_projects_dict[str(project).strip()] = [sheet.cell(row, col).value for col in range(2, 5)]
        row += 1

    wb.close()

    return dict(sorted(all_projects_dict.items()))


def calculate_data_projects(finance_projects_dict: dict, 
                            finance_employees_dict: dict, 
                            employees: list,
                            year: float = date.today().year):
    type_salary = TYPES_FINANCES[0]

    # суммарные траты проектов на з/п по месяцам
    sum_expenses_projects = {}
    for employee in employees:
        name_employee = employee.get_full_name()
        work_hours_dict = employee.get_work_hours()
        # сумма затраченных часов сотрудника по месяцам
        sum_hours = {}
        for date, project_hours_dict in work_hours_dict.items():
            month = MONTHS[date.month-1]
            for hours in project_hours_dict.get_date_time_spent().values():
                sum_hours[month] = sum_hours.get(month, 0) + hours
        
        for date, project_hours_dict in work_hours_dict.items():
            month = MONTHS[date.month-1]
            try:
                month_salary = finance_employees_dict[name_employee][type_salary][month]
            except KeyError:
                messagebox.showerror("Ошибка", f'Сотрудник "{name_employee}" отсутствует в файле с заполненными З/П и премиями!')
                raise KeyError(f'Сотрудник {name_employee} отсутствует в файле с заполненными З/П и премиями!')
            if month_salary is None:
                month_salary = 0
            for project, hours in project_hours_dict.get_date_time_spent().items():
                expense = month_salary * (hours / sum_hours[month])
                sum_expenses_projects[project] = sum_expenses_projects.get(project, {})
                sum_expenses_projects[project][month] = sum_expenses_projects[project].get(month, 0) + expense

    bonuses_projects, beginning_year_balance = {}, {}

    for project, incomes_project in finance_projects_dict.items():
        beginning_year_balance[project] = incomes_project[f'Бюджет на 01.01.{year} г.']
        bonuses_projects[project] = {}
        for month in MONTHS:
            bonuses_projects[project][month] = incomes_project[f'{month} - {TYPES_FINANCES[1]}']

    return sum_expenses_projects, bonuses_projects, beginning_year_balance


def create_data_projects_file(filename: str, 
                              column_names_projects: dict,
                              sorted_all_projects_list: list,
                              beginning_year_balances: dict,
                              sum_expenses_projects: dict,
                              bonuses_projects: dict,
                              year: float = date.today().year, 
                              sheet_name: str = "Траты"):
    
    sheet_title = full_sheet_title(sheet_name, year)
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
    
    if sheet_title in wb.sheetnames:
        messagebox.showwarning("Внимание!", f'Лист с итоговыми расчётами за {year} год уже существует! Удалите лист или измените имя листа в файле!')  
    else:
        sheet = wb.create_sheet(sheet_title)
    sheet.freeze_panes = 'D2'

    name_columns = column_names_projects
    name_columns.append(f'Бюджет на 01.01.{year} г.')
    for month in MONTHS:
        for finance in TYPES_FINANCES:
            name_columns.append(f'{month} - {finance}')
    name_columns.extend(['Суммарные затраты на З/П', 
                         'Суммарные затраты на премии',
                         'Остаток бюджета'])

    for col in range(1, len(name_columns)+1):
        sheet.cell(1, col).value = name_columns[col-1]
        sheet.column_dimensions[num_col_to_char_excel(col)].width = 22
        sheet.cell(1, col).alignment = CENTER_ALIGNMENT_WITH_WRAP

    row = 2
    for project, other_names_projects in sorted_all_projects_list.items():
        
        sheet.cell(row, 1).value = project
        sheet.cell(row, 1).alignment = CENTER_ALIGNMENT_WITH_WRAP
        col = 2
        for name_project in other_names_projects:
            sheet.cell(row, col).value = name_project
            sheet.cell(row, col).alignment = CENTER_ALIGNMENT_WITH_WRAP
            col += 1
        
        try:
            year_expenses_project = sum_expenses_projects[project]
        except KeyError:
            sheet.cell(row, 1).fill = PatternFill(start_color='FFA500', 
                                                    end_color='FFA500', 
                                                    fill_type='solid')
            year_expenses_project = {}

        sheet.cell(row, col).value = beginning_year_balances[project]
        sheet.cell(row, col).alignment = CENTER_ALIGNMENT_WITHOUT_WRAP
        cell_begin_balance = row_col_to_cell_excel(row, col)
        col += 1
        cells_salary, cells_bonuses = [], []
        for month in MONTHS:
            sheet.cell(row, col).value = year_expenses_project.get(month, 0)
            sheet.cell(row, col).alignment = CENTER_ALIGNMENT_WITHOUT_WRAP
            cells_salary.append(row_col_to_cell_excel(row, col))
            col += 1
            sheet.cell(row, col).value = bonuses_projects[project].get(month, 0)
            sheet.cell(row, col).alignment = CENTER_ALIGNMENT_WITHOUT_WRAP
            cells_bonuses.append(row_col_to_cell_excel(row, col))
            col += 1
        
        sheet.cell(row, col).value = f'={"+".join(cells_salary)}'
        sheet.cell(row, col).alignment = CENTER_ALIGNMENT_WITHOUT_WRAP
        cell_total_expenses_salary = row_col_to_cell_excel(row, col)
        col += 1

        sheet.cell(row, col).value = f'={"+".join(cells_bonuses)}'
        sheet.cell(row, col).alignment = CENTER_ALIGNMENT_WITHOUT_WRAP
        cell_total_expenses_bonuses = row_col_to_cell_excel(row, col)
        col += 1

        sheet.cell(row, col).value = f'={cell_begin_balance}-({cell_total_expenses_salary}+{cell_total_expenses_bonuses})'
        sheet.cell(row, col).alignment = CENTER_ALIGNMENT_WITHOUT_WRAP

        row += 1

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    redFill = PatternFill(start_color='FF0000',
                          end_color='FF0000',
                          fill_type='solid')
    greenFill = PatternFill(start_color='00FF00',
                          end_color='00FF00',
                          fill_type='solid')
    
    first_cell = row_col_to_cell_excel(2, col)
    last_cell = row_col_to_cell_excel(row-1, col)
    
    sheet.conditional_formatting.add(f'{first_cell}:{last_cell}', 
                                     CellIsRule(operator='greaterThanOrEqual', formula=['0'], stopIfTrue=True, fill=greenFill))
    sheet.conditional_formatting.add(f'{first_cell}:{last_cell}', 
                                     CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=redFill))
    wb.save(filename)
    wb.close()  


def putting_resulting_formuls_in_finance_file(filename: str, 
                                              year: float = date.today().year):
    if not os.path.exists(filename):
        messagebox.showwarrning('Внимание!', 'Указанный файл отсутствует в директории!')
    wb = load_workbook(filename)
    SHEETNAME_FINANCE_EMPLOYEES = full_sheet_title('Финансы сотрудников', year)
    SHEETNAME_FINANCE_PROJECTS = full_sheet_title('Бюджеты проектов', year)

    cell_bonuses_employees = list()
    for namesheet in [SHEETNAME_FINANCE_EMPLOYEES, SHEETNAME_FINANCE_PROJECTS]:
        sheet = wb[namesheet]
        last_row = 1
        while True:
            if sheet.cell(last_row, 1).value is None:
                break
            last_row += 1
        col = 1
        if namesheet == SHEETNAME_FINANCE_EMPLOYEES:
            while True:
                column_name = sheet.cell(1, col).value
                if column_name is None:
                    break
                elif TYPES_FINANCES[0] in column_name or TYPES_FINANCES[1] in column_name:
                    sheet.cell(last_row, col).value = f'=SUM({row_col_to_cell_excel(2, col)}:{row_col_to_cell_excel(last_row-1, col)})'
                    sheet.cell(last_row, col).alignment = CENTER_ALIGNMENT_WITHOUT_WRAP
                    if TYPES_FINANCES[1] in column_name:
                        cell_bonuses_employees.append([last_row, col])
                col += 1
        elif namesheet == SHEETNAME_FINANCE_PROJECTS:
            num_el_cell = 0
            while True:
                column_name = sheet.cell(1, col).value
                if column_name is None:
                    break
                elif TYPES_FINANCES[1] in column_name:
                    sheet.cell(last_row, col).value = f'=SUM({row_col_to_cell_excel(2, col)}:{row_col_to_cell_excel(last_row-1, col)})'
                    sheet.cell(last_row, col).alignment = CENTER_ALIGNMENT_WITHOUT_WRAP

                    sheet.cell(last_row+1, col).value = f"='{SHEETNAME_FINANCE_EMPLOYEES}'!{row_col_to_cell_excel(*cell_bonuses_employees[num_el_cell])}-{row_col_to_cell_excel(last_row, col)}"
                    sheet.cell(last_row+1, col).alignment = CENTER_ALIGNMENT_WITHOUT_WRAP
                    num_el_cell += 1
                col += 1
        else:
            continue
            
    wb.save(filename)
    wb.close()


def parse_year_report():
    global column_names_projects
    global employees
    global sorted_projects_list
    global sorted_employee_position_list
    global projects_names_list

    filename_year_report = file_year_report.get()
    if filename_year_report == '':
        messagebox.showerror('Ошибка', f'Файл отчёта не был выбран!')
    else:
        year = selected_year.get()
        column_names_projects = parse_column_names_projects_to_list(filename_year_report)
        employees, sorted_projects_list = parse_info_employees_projects_to_lists(filename_year_report, 
                                                                                 year)

        employee_position_list = [[employee.get_full_name(), employee.get_position()] for employee in employees]
        sorted_employee_position_list = sorted(employee_position_list, key=lambda x: x[0])
        
        projects_names_list = [(num_project, *names_project) for num_project, names_project in sorted_projects_list.items()]
        messagebox.showinfo('Информация', 'Данные из выгрузки годового отчёта из Redmine успешно прочитаны')


flag_finance_employees_projects = False


def create_empty_file():
    # создание пустого файла для заполнения з/п и премий сотрудников и бюджетов проектов
    global flag_finance_employees_projects

    filename_finance_employees_projects = entry_fin_empl_proj.get()
    if filename_finance_employees_projects == '':
        messagebox.showerror("Ошибка", f'Введите наименование файла!')
    else:
        try:
            year = selected_year.get()
            create_empty_finance_employees_sheet(filename_finance_employees_projects, 
                                                 sorted_employee_position_list, 
                                                 year)
            create_empty_budget_projects_sheet(filename_finance_employees_projects, 
                                               projects_names_list, 
                                               column_names_projects, 
                                               year)
            flag_finance_employees_projects = True
            messagebox.showinfo('Информация', 'Успешно создан пустой файл для заполнения З/П и премий')
        except NameError:
            messagebox.showerror("Ошибка", f'Сначала прочтите данные из файла годового отчёта!')
        

def add_new_employees_and_projects():
    global flag_finance_employees_projects

    filename_update_empl_projs = file_update_empl_proj.get()
    if filename_update_empl_projs == '':
        messagebox.showerror("Ошибка", f'Введите наименование файла с З/П и премиями сотрудников и проектов!')
    else:
        try:
            year = selected_year.get()
            add_new_employees(filename_update_empl_projs, 
                              sorted_employee_position_list, 
                              year)
            add_new_projects(filename_update_empl_projs, 
                             projects_names_list, 
                             column_names_projects, 
                             year)
            messagebox.showinfo('Информация', 'Файл успешно обновлён')
        except:
            messagebox.showerror("Ошибка", f'Проверьте не открыт ли файл и прочитаны ли данные из файла годового отчёта!')


def create_final_data_projects_file():
    filename_filled_empl_proj = file_filled_empl_proj.get()
    filename_final_data = entry_final_data_file.get()
    year = selected_year.get()
    if filename_filled_empl_proj == '':
        messagebox.showerror("Ошибка", f'Выберите файл с заполненными З/П и премиями сотрудников и проектов!')
    elif filename_final_data == '':
        messagebox.showerror("Ошибка", f'Заполните имя итогового файла!')
    # elif not flag_finance_employees_projects:
    #     messagebox.showinfo("Ошибка", f'Сначала создайте/апдейтните файл для заполнения З/П и премий!')
    else:
        #считывание з/п и премий сотрудников и бюджетов проектов из заполненного файла
        completed_finance_employees_dict = parse_completed_finance_employees(filename_filled_empl_proj, 
                                                                             year)
        completed_finance_projects_dict = parse_completed_finance_projects(filename_filled_empl_proj, 
                                                                           year)
        
        # считывание полного списка проектов из заполненного файла з/п и премий
        sorted_all_projects_list = parse_all_projects(filename_filled_empl_proj, year)
        sum_expenses_projects, bonuses_projects, beginning_year_balances = calculate_data_projects(completed_finance_projects_dict,
                                                                                                   completed_finance_employees_dict, 
                                                                                                   employees, 
                                                                                                   year)
        # создание файла трат з/п по проектам за год
        create_data_projects_file(filename_final_data, 
                                column_names_projects, 
                                sorted_all_projects_list, 
                                beginning_year_balances, 
                                sum_expenses_projects,
                                bonuses_projects,
                                year)
        putting_resulting_formuls_in_finance_file(filename_filled_empl_proj, 
                                                  year)
        messagebox.showinfo('Информация', 'Готово!\nСоздан итоговый файл данных')


def view_files_excel_list():
    all_files = os.listdir(os.getcwd())
    files_excel_list = list(file for file in all_files if 'xls' in file and '~' not in file)
    return files_excel_list


if __name__ == '__main__':
    # сегодняшний год
    curr_year = list(map(int, str(date.today()).split('-')))[0]

    # цвет фона окна
    background_color = '#4BA7E9'
    root = Tk()
    root.title("Программа обработки данных по сотрудникам и проектам")
    root.geometry("550x690")
    root.configure(background=background_color)

    label = Label(text='Выберите год, за который хотите обработать данные:', 
                   background=background_color)
    label.pack(anchor='w', pady=5)
    selected_year = ttk.Combobox(root, values=AVAILABLE_YEARS, width=8)
    selected_year.current(AVAILABLE_YEARS.index(curr_year))
    selected_year.pack(anchor='c', pady=5)

    label = Label(text='1. Считывание файла выгрузки "Отчёт за текущий год" из Redmine', 
                   background=background_color)
    label.pack(anchor='w', pady=5)

    label = Label(text="Выберите файл выгрузки отчета за год из RedMine:", background=background_color)
    label.pack(anchor="c", pady=5) 
    file_year_report = ttk.Combobox(root, values=view_files_excel_list())
    file_year_report.pack(fill=X, padx=90, pady=5)
    btn_1 = Button(text="Считать данные из файла годового отчёта", 
                   bg='#18C0AF', 
                   font=('uni sans', 12, 'bold'), 
                   command=parse_year_report)
    btn_1.pack(anchor='c', pady=5)

    label = Label(text='2. Создание/апдейт файла для заполнения З/П и премий сотрудников и бюджетов проектов', 
                   background=background_color)
    label.pack(anchor='w', pady=10)
    
    label = Label(text='2.1. Либо укажите наименование для пустого файла с З/П и премиями:', 
                  background=background_color)
    label.pack(anchor='c', pady=5)

    entry_fin_empl_proj = ttk.Entry(width=50)
    entry_fin_empl_proj.pack(anchor='c', pady=5)
    entry_fin_empl_proj.insert(0, 'Finance employees and projects.xlsx')

    btn_2 = Button(text="Создать пустой файл для заполнения", 
                   bg='#18C0AF', 
                   font=('uni sans', 12, 'bold'), 
                   command=create_empty_file)
    btn_2.pack(anchor='c', pady=5)

    label = Label(text='2.2. Либо выберите уже существующий файл для записи новых сотрудников и проектов:', 
                  background=background_color)
    label.pack(anchor='c', pady=5)

    file_update_empl_proj = ttk.Combobox(root, values=view_files_excel_list())
    file_update_empl_proj.pack(fill=X, padx=90, pady=5)

    btn_3 = Button(text="Вписать новых сотрудников и проекты в файл с З/П и премиями", 
                   bg='#18C0AF', 
                   font=('uni sans', 12, 'bold'), 
                   command=add_new_employees_and_projects)
    btn_3.pack(anchor='c', pady=5)

    label = Label(text='3. Генерация итогового файла данных по сотрудникам и проектам за год', 
                  background=background_color)
    label.pack(anchor='w', pady=5)

    label = Label(text='Выберите заполненный файл с З/П и премиями сотрудников и проектов', 
                  background=background_color)
    label.pack(anchor='c', pady=5)

    file_filled_empl_proj = ttk.Combobox(root, values=view_files_excel_list())
    file_filled_empl_proj.pack(fill=X, padx=90, pady=5)

    label = Label(text='Введите наименование итогового файла данных по сострудникам и проектам:', 
                  background=background_color)
    label.pack(anchor='c', pady=5)

    entry_final_data_file = ttk.Entry(width=50)
    entry_final_data_file.pack(anchor='c', pady=5)
    entry_final_data_file.insert(0, 'Data of finance projects for year.xlsx')

    btn_4 = Button(text="Сгенерировать итоговый файл данных", 
                   bg='#18C0AF', 
                   font=('uni sans', 12, 'bold'), 
                   command=create_final_data_projects_file)
    btn_4.pack(anchor='c', pady=5)

    btn_5 = Button(text='Закрыть программу',
                   bg='#FF0000', 
                   font=('uni sans', 12, 'bold'), 
                   command=root.destroy)
    btn_5.pack(anchor='c', pady=5)

    root.mainloop()
